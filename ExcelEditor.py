import openpyxl


class ExcelEditor():
    def __init__(self, filename=None, sheetname=None):
        self.wb = None
        self.initialize_excel(filename, sheetname)

    def initialize_excel(self, filename, sheetname):
        if filename:
            self.open_excel(filename)
        elif sheetname:
            self.create_excel(sheetname)

    def create_excel(self, sheetname):
        self.wb = openpyxl.Workbook()
        self.add_new_sheet(sheetname)

    def open_excel(self, filename):
        self.wb = openpyxl.load_workbook(filename)

    def add_new_sheet(self, sheetname):
        self.wb.create_sheet(title=sheetname)

    def remove_sheet(self, sheetname):
        sheet = self.wb.get_sheet_by_name(sheetname)
        self.wb.remove_sheet(sheet)

    def get_sheet_names(self):
        return self.wb.get_sheet_names()

    def get_column_count(self, sheetname):
        return self.wb[sheetname].max_column

    def get_row_count(self, sheetname):
        return self.wb[sheetname].max_row

    def get_row(self, sheetname, row_value):
        return list(map(lambda x: x.value, list(self.wb[sheetname].rows)[row_value - 1]))

    def read_cell_data_by_coordinates(self, sheetname, row, column):
        return self.wb[sheetname].cell(row=int(row), column=int(column)).value

    def write_data_by_coordinates(self, sheetname, row, column, value):
        self.wb[sheetname].cell(row=int(row), column=int(column)).value = value

    def save_excel(self, filename):
        self.wb.save(filename)
